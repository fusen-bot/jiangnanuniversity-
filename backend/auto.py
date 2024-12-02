import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import os

def read_excel(data):
    """
    读取Excel文件，支持.xls和.xlsx格式
    """
    try:
        if data.endswith('.xlsx'):
            return pd.read_excel(data, engine='openpyxl')
        elif data.endswith('.xls'):
            return pd.read_excel(data, engine='xlrd')
        else:
            raise ValueError(f"不支持的文件格式: {data}")
    except Exception as e:
        print(f"读取文件 {data} 时出错: {e}")
        return None

def load_employee_data(employee_file, retired_file):
    employee_df = read_excel(employee_file)
    retired_df = read_excel(retired_file)
    if employee_df is None or retired_df is None:
        print("错误：无法读取员工数据文件")
        return None
    combined_df = pd.concat([employee_df, retired_df])
    combined_df['工号'] = pd.to_numeric(combined_df['工号'], errors='coerce').fillna(0).astype(int)
    return combined_df

def query_by_employee_id(employee_data, employee_id_str):
    # 处理多个工号查询
    employee_ids = [id.strip() for id in employee_id_str.split('、')]
    all_results = pd.DataFrame()
    
    for emp_id in employee_ids:
        try:
            emp_id = int(emp_id)  # 确保工号是整数
            result = employee_data[employee_data['工号'] == emp_id]
            all_results = pd.concat([all_results, result])
        except ValueError:
            print(f"警告：无效的工号格式 '{emp_id}'")
    
    if all_results.empty:
        print(f"未找到工号为 {employee_id_str} 的员工")
    else:
        # 检查是否有重复工号
        if len(all_results) > len(employee_ids):
            print("注意：某些工号对应多个员工记录")
        print("\n查询结果：")
        print(all_results[['姓名', '工号', '部门']].to_string(index=False, float_format='{:.0f}'.format))

def query_by_name(employee_data, name_str):
    # 处理多个姓名查询
    names = [name.strip() for name in name_str.split('、')]
    all_results = pd.DataFrame()
    
    for name in names:
        result = employee_data[employee_data['姓名'] == name]
        all_results = pd.concat([all_results, result])
    
    if all_results.empty:
        print(f"未找到姓名为 {name_str} 的员工")
    else:
        # 检查是否有重名情况
        name_counts = all_results['姓名'].value_counts()
        duplicate_names = name_counts[name_counts > 1]
        if not duplicate_names.empty:
            print("\n注意：以下姓名存在多条记录：")
            for name, count in duplicate_names.items():
                print(f"- {name}: {count}条记录")
        
        print("\n查询结果：")
        print(all_results[['姓名', '工号', '部门']].to_string(index=False, float_format='{:.0f}'.format))



def is_internal(unit, address):
    """
    判断审稿人是否为校内人员
    """
    keywords = ['江南大学', '蠡湖大道', '1800号']
    return any(keyword in str(unit) or keyword in str(address) for keyword in keywords)

def match_employee_id(name, employee_df):
    """
    匹配员工工号，并检查是否有重复
    """
    matches = employee_df[employee_df['姓名'] == name]
    if len(matches) == 0:
        return None, False
    elif len(matches) == 1:
        return matches.iloc[0]['工号'], False
    else:
        return matches.iloc[0]['工号'], True  # 返回第一个匹配的工号，并标记为重复

def get_target_month():
    """
    获取用户输入的目标月份，并转换 YYYY-MM 格式
    """
    while True:
        month_input = input("请输入目标月份（1-12）: ")
        try:
            month = int(month_input)
            if 1 <= month <= 12:
                current_year = datetime.now().year
                return f"{current_year}-{month:02d}"
            else:
                print("错误：月份必须在 1 到 12 之间")
        except ValueError:
            print("错误：请输入有效的数字")

def process_review_data(review_file, re_review_file, employee_file, retired_file, target_month):
    """
    处理审稿数据的主函数
    """
    # 读取文件
    review_df = read_excel(review_file)
    re_review_df = read_excel(re_review_file)
    employee_df = read_excel(employee_file)
    retired_df = read_excel(retired_file) 

    review_df['稿件编号'] = review_df['稿件编号'].astype(str)
    re_review_df['稿件编号'] = re_review_df['稿件编号'].astype(str)


    if review_df is None or re_review_df is None or employee_df is None or retired_df is None:
        print("错误：一个或多个必要的文件无法读取。请检查文件路径和格式。")
        return None
    
    # 转换日期列
    review_df['审回时间'] = pd.to_datetime(review_df['审回时间'])
    re_review_df['审回时间'] = pd.to_datetime(re_review_df['审回时间'])

    # 筛选目标月份数据
    target_date = datetime.strptime(target_month, "%Y-%m")
    start_date = target_date.replace(day=1)
    end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    next_month_end = end_date + timedelta(days=31)

    review_df = review_df[(review_df['审回时间'] >= start_date) & (review_df['审回时间'] <= end_date)]
    re_review_df = re_review_df[(re_review_df['审回时间'] >= start_date) & (re_review_df['审回时间'] <= next_month_end)]

    # 初始化费用
    review_df['审稿费用'] = 150
    review_df['复审'] = False
    review_df['复审月份'] = ''

    # 处理复审数据
    for _, row in re_review_df.iterrows():
        mask = (review_df['稿件编号'] == row['稿件编号'])
        if mask.any():
            review_df.loc[mask, '审稿费用'] = review_df.loc[mask, '审稿费用'].apply(lambda x: min(x + 50, 200))
            review_df.loc[mask, '复审'] = True
            current_month = row['审回时间'].strftime('%Y-%m')
            review_df.loc[mask, '复审月份'] += f"{current_month}," if review_df.loc[mask, '复审月份'].values[0] else current_month
        else:
            print(f"警告：复审记录无配的评审记录 - 稿件编号: {row['稿件编号']}")

    #复审无匹配统计待增加
    

    # 判断校内/校外
    review_df['校内人员'] = review_df.apply(lambda row: is_internal(row['审稿人单位'], row['审稿人地址']), axis=1)

    # 工号匹配
    def match_employee_info(name, employee_df, retired_df):
        matches = employee_df[employee_df['姓名'] == name]
        if len(matches) == 0:
            # 在退休人员中查找
            retired_matches = retired_df[retired_df['姓名'] == name]
            if len(retired_matches) == 0:
                return None, None, False, 'not_found'
            elif len(retired_matches) == 1:
                return retired_matches.iloc[0]['工号'], retired_matches.iloc[0]['部门'], False, 'retired'
            else:
                return retired_matches.iloc[0]['工号'], retired_matches.iloc[0]['部门'], True, 'retired'
        elif len(matches) == 1:
            return matches.iloc[0]['工号'], matches.iloc[0]['部门'], False, 'active'
        else:
            return matches.iloc[0]['工号'], matches.iloc[0]['部门'], True, 'active'

    review_df['工号'] = ''
    review_df['部门'] = ''
    review_df['工号重复'] = False
    review_df['人员状态'] = ''
    for idx, row in review_df[review_df['校内人员']].iterrows():
        emp_id, department, is_duplicate, status = match_employee_info(row['审稿人姓名'], employee_df, retired_df)
        review_df.at[idx, '工号'] = emp_id if emp_id else ''
        review_df.at[idx, '部门'] = department if department else ''
        review_df.at[idx, '工号重复'] = is_duplicate
        review_df.at[idx, '人员状态'] = status

    print("审稿数据列名:", review_df.columns.tolist())  # 保留这行来帮助调试

    # 检查校外专家的银行账户信息
    review_df['银行账户缺失'] = False
    # 使用确切的列名
    bank_account_column = '银行账号'
    bank_name_column = '开户银行'
    
    if bank_account_column in review_df.columns and bank_name_column in review_df.columns:
        # 检查银行账号和开户银行是否都为空
        review_df.loc[~review_df['校内人员'], '银行账户缺失'] = (
            review_df.loc[~review_df['校内人员'], bank_account_column].isna() & 
            review_df.loc[~review_df['校内人员'], bank_name_column].isna()
        )
        print(f"已检查银行账户信息。列名: {bank_account_column} 和 {bank_name_column}")
    else:
        missing_columns = []
        if bank_account_column not in review_df.columns:
            missing_columns.append(bank_account_column)
        if bank_name_column not in review_df.columns:
            missing_columns.append(bank_name_column)
        print(f"警告：未找到以下银行账户信息列：{', '.join(missing_columns)}。无法完全检查银行账户信息。")
        print("可用的列名：", review_df.columns.tolist())

    return review_df

def highlight_cells(worksheet, column, condition_column, condition_value, color):
    """
    在Excel中高亮显示特定单元格
    """
    for cell in worksheet[column]:
        if worksheet[condition_column + str(cell.row)].value == condition_value:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def save_to_excel(data, target_month):
    """保存数据到Excel文件并在数据下方添加统计报表"""
    output_file = os.path.join(OUTPUT_PATH, f'review_fee_{target_month}.xlsx')
    
    # 创建DataFrame并保存到Excel
    df = pd.DataFrame(data)
    
    # 创建ExcelWriter对象
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 写入主数据
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # 获取工作簿和工作表对象
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # 获取数据最后一行的位置
        last_row = len(df) + 2  # 加2是因为标题行和从1开始计数
        
        # 在数据下方添加统计报表
        add_detailed_statistics(worksheet, df)
    
    return output_file

def add_detailed_statistics(worksheet, df):
    """生成简化版审稿费统计报表"""
    bold_font = Font(bold=True)
    last_row = worksheet.max_row + 2

    # 添加标题
    worksheet.cell(row=last_row, column=1, value="评审专家统计").font = bold_font
    last_row += 1

    # 校内在职专家统计
    worksheet.cell(row=last_row, column=1, value="校内在职专家").font = bold_font
    worksheet.cell(row=last_row, column=2, value="工号").font = bold_font
    worksheet.cell(row=last_row, column=3, value="金额").font = bold_font
    last_row += 1

    internal_active = df[(df['校内人员']) & (df['人员状态'] == 'active')]
    internal_active_grouped = internal_active.groupby(['审稿人姓名', '工号'])['审稿费金额'].sum().reset_index()
    internal_active_grouped = internal_active_grouped.sort_values('审稿费金额', ascending=False)

    for _, row in internal_active_grouped.iterrows():
        worksheet.cell(row=last_row, column=1, value=row['审稿人姓名'])
        worksheet.cell(row=last_row, column=2, value=row['工号'])
        worksheet.cell(row=last_row, column=3, value=row['审稿费金额'])
        last_row += 1

    # 校内在职总金额
    worksheet.cell(row=last_row, column=1, value="校内在职总金额").font = bold_font
    worksheet.cell(row=last_row, column=3, value=internal_active_grouped['审稿费金额'].sum()).font = bold_font
    last_row += 2

    # 校内退休专家统计
    worksheet.cell(row=last_row, column=1, value="校内退休专家").font = bold_font
    worksheet.cell(row=last_row, column=2, value="工号").font = bold_font
    worksheet.cell(row=last_row, column=3, value="金额").font = bold_font
    last_row += 1

    internal_retired = df[(df['校内人员']) & (df['人员状态'] == 'retired')]
    internal_retired_grouped = internal_retired.groupby(['审稿人姓名', '工号'])['审稿费金额'].sum().reset_index()
    internal_retired_grouped = internal_retired_grouped.sort_values('审稿费金额', ascending=False)

    for _, row in internal_retired_grouped.iterrows():
        worksheet.cell(row=last_row, column=1, value=row['审稿人姓名'])
        worksheet.cell(row=last_row, column=2, value=row['工号'])
        worksheet.cell(row=last_row, column=3, value=row['审稿费金额'])
        last_row += 1

    # 校内退休总金额
    worksheet.cell(row=last_row, column=1, value="校内退休总金额").font = bold_font
    worksheet.cell(row=last_row, column=3, value=internal_retired_grouped['审稿费金额'].sum()).font = bold_font
    last_row += 2

    # 校外专家统计
    worksheet.cell(row=last_row, column=1, value="校外专家").font = bold_font
    worksheet.cell(row=last_row, column=2, value="身份证号").font = bold_font
    worksheet.cell(row=last_row, column=3, value="金额").font = bold_font
    last_row += 1

    external = df[~df['校内人员']]
    external_grouped = external.groupby(['审稿人姓名', '审稿人身份证号'])['审稿费金额'].sum().reset_index()
    external_grouped = external_grouped.sort_values('审稿费金额', ascending=False)

    for _, row in external_grouped.iterrows():
        worksheet.cell(row=last_row, column=1, value=row['审稿人姓名'])
        worksheet.cell(row=last_row, column=2, value=row['审稿人身份证号'] if pd.notna(row['审稿人身份证号']) else '')
        worksheet.cell(row=last_row, column=3, value=row['审稿费金额'])
        last_row += 1

    # 校外总金额
    worksheet.cell(row=last_row, column=1, value="校外总金额").font = bold_font
    worksheet.cell(row=last_row, column=3, value=external_grouped['审稿费金额'].sum()).font = bold_font
    last_row += 2

    # 总计
    worksheet.cell(row=last_row, column=1, value="总计").font = bold_font
    total = df['审稿费金额'].sum()
    worksheet.cell(row=last_row, column=3, value=total).font = bold_font

    # 设置金额列的数字格式
    for row in range(worksheet.max_row - worksheet.min_row + 1):
        cell = worksheet.cell(row=row + 1, column=3)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

def query_by_manuscript_id_or_reviewer(review_file, re_review_file, query):
    review_df = read_excel(review_file)
    re_review_df = read_excel(re_review_file)

    if review_df is None or re_review_df is None:
        print("错误：无法读取评审表或复审表文件")
        return

    review_df['稿件编号'] = review_df['稿件编号'].astype(str)
    re_review_df['稿件编号'] = re_review_df['稿件编号'].astype(str)

    # 查询评审表
    review_result = review_df[(review_df['稿件编号'] == query) | (review_df['审稿人姓名'] == query)]
    
    if not review_result.empty:
        print("\n在评审表中找到以下记录：")
        for _, row in review_result.iterrows():
            print(f"稿件编号: {row['稿件编号']}, 审稿人: {row['审稿人姓名']}, 审回时间: {row['审回时间'].strftime('%Y-%m-%d')}")
    else:
        print("在评审表中未找到相关记录")

    # 查询复审表
    re_review_result = re_review_df[(re_review_df['稿件编号'] == query) | (re_review_df['审稿人姓名'] == query)]
    
    if not re_review_result.empty:
        print("\n在复审表中到以下记录：")
        for _, row in re_review_result.iterrows():
            print(f"稿件编号: {row['稿件编号']}, 审稿人: {row['审稿人姓名']}, 审回时间: {row['审回时间'].strftime('%Y-%m-%d')}")
    else:
        print("在复审表中未找到相关记录")

def process_and_save_data(review_file, re_review_file, employee_file, retired_file, target_month):
    """处理数据并保存到Excel"""
    # 处理数据
    data = process_review_data(review_file, re_review_file, employee_file, retired_file, target_month)
    if data is None:
        return None
    
    # 保存到Excel并添加统计
    output_file = save_to_excel(data, target_month)
    return output_file