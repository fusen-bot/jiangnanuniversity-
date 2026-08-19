from flask import Flask, request, jsonify, send_from_directory
from data_processing import process_review_data, query_by_employee_id, query_by_name, query_by_manuscript_id_or_reviewer
from database import load_employee_data
import json
import logging
import os
import platform
import subprocess
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from spark_chat_interactive import SparkChatBot
from webpq import scrape_and_process_data
from werkzeug.utils import secure_filename

from paper_author_extractor import extract_author_info, create_excel_report

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

app = Flask(__name__, static_folder='../frontend')


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = Path(os.getenv('APP_DATA_DIR', PROJECT_ROOT / 'data')).expanduser()
OUTPUT_PATH = Path(os.getenv('APP_OUTPUT_DIR', PROJECT_ROOT / 'output')).expanduser()
PAGE_FEE_FILE = Path(os.getenv('PAGE_FEE_FILE', DATA_DIR / 'page_fee.xlsx')).expanduser()
NOTES_FILE = Path(os.getenv('APP_NOTES_FILE', DATA_DIR / 'notes.json')).expanduser()
UPLOAD_FOLDER = Path(os.getenv('APP_UPLOAD_DIR', PROJECT_ROOT / 'uploads')).expanduser()
WORKSPACE_DIR = Path(os.getenv('APP_WORKSPACE_DIR', PROJECT_ROOT)).expanduser()
PROGRAM_DIR = Path(os.getenv('APP_PROGRAM_DIR', PROJECT_ROOT / 'backend')).expanduser()
ALLOWED_EXTENSIONS = {'pdf'}

review_file = str(DATA_DIR / '评审.xls')
re_review_file = str(DATA_DIR / '复审.xls')
employee_file = str(DATA_DIR / '在职员工.xlsx')
retired_file = str(DATA_DIR / '退休员工.xlsx')

OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
NOTES_FILE.parent.mkdir(parents=True, exist_ok=True)

logging.info('应用路径已初始化')

try:
    logging.info('开始加载员工数据')
    employee_data = load_employee_data(employee_file, retired_file)
    logging.info('员工数据加载完成')
except FileNotFoundError:
    logging.exception('员工数据文件缺失')
    employee_data = []
except Exception:
    logging.exception('加载员工数据失败')
    raise

spark_chatbot = SparkChatBot()


def _server_error(message: str):
    return jsonify({'error': message}), 500


def _open_directory(directory: Path):
    if platform.system() == 'Darwin':
        subprocess.run(['open', str(directory)], check=True)
    elif platform.system() == 'Windows':
        subprocess.run(['explorer', str(directory)], check=True)
    else:
        subprocess.run(['xdg-open', str(directory)], check=True)


@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')


@app.route('/process_review', methods=['POST'])
def process_review():
    target_month = request.json.get('target_month')
    if not target_month:
        return jsonify({'error': '缺少 target_month 参数'}), 400

    try:
        result = process_review_data(review_file, re_review_file, employee_file, retired_file, target_month)
        if isinstance(result, list) and result:
            df = pd.DataFrame(result)
            output_file = OUTPUT_PATH / f'processed_review_{target_month}.xlsx'
            output_file.parent.mkdir(parents=True, exist_ok=True)
            df.to_excel(output_file, index=False)
            return jsonify({'message': '数据处理成功', 'file': output_file.name}), 200
        return jsonify({'error': '没有找到符合条件的数据'}), 404
    except Exception:
        logging.exception('处理审稿数据失败')
        return jsonify({'error': '处理失败'}), 400


@app.route('/query_employee', methods=['GET'])
def query_employee():
    employee_id = request.args.get('employee_id')
    name = request.args.get('name')

    try:
        if employee_id:
            result = query_by_employee_id(employee_data, employee_id)
        elif name:
            result = query_by_name(employee_data, name)
        else:
            return jsonify({'error': '请提供员工ID或姓名'}), 400

        if isinstance(result, dict) and 'message' in result:
            return jsonify({'error': result['message']}), 404

        return jsonify({'type': 'employee', 'data': result})
    except Exception:
        logging.exception('查询员工信息失败')
        return _server_error('查询失败')


@app.route('/query_manuscript', methods=['GET'])
def query_manuscript():
    query = request.args.get('query')
    result = query_by_manuscript_id_or_reviewer(review_file, re_review_file, query)
    return jsonify({'type': 'manuscript', 'data': result})


@app.route('/get_page_fee_data', methods=['GET'])
def get_page_fee_data():
    try:
        wb = load_workbook(PAGE_FEE_FILE)
        ws = wb.active

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                status_value = row[10] if len(row) > 10 else ''
                statuses = set(status_value.split(', ')) if status_value and status_value != '已完成' else {'录用', '发票'}
                data.append({
                    '备注': row[0] or '',
                    '稿件编号': str(row[1]),
                    '核销号': row[2] or '',
                    '财务备注': row[5] or '' if len(row) > 5 else '',
                    '税号': row[6] or '' if len(row) > 6 else '',
                    '发票抬头': row[4] or '' if len(row) > 4 else '',
                    '邮箱': row[7] or '' if len(row) > 7 else '',
                    '录用': '录用' in statuses,
                    '发票': '发票' in statuses
                })

        data.reverse()
        return jsonify(data)
    except Exception:
        logging.exception('读取版面费数据失败')
        return _server_error('读取版面费数据失败')


@app.route('/update_status', methods=['POST'])
def update_status():
    try:
        data = request.json
        manuscript_number = data.get('稿件编号')
        status_type = data.get('statusType')
        is_checked = data.get('isChecked')

        if not manuscript_number or status_type not in ['录用', '发票']:
            return jsonify({'error': '无效的请求数据'}), 400

        wb = load_workbook(PAGE_FEE_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[1].value == manuscript_number:
                status_cell = row[10]
                current_status = status_cell.value or ''
                statuses = set(current_status.split(', ')) if current_status else set()

                if is_checked:
                    statuses.add(status_type)
                else:
                    statuses.discard(status_type)

                if '录用' in statuses and '发票' in statuses:
                    status_cell.value = '已完成'
                else:
                    status_cell.value = ', '.join(sorted(statuses))
                break

        wb.save(PAGE_FEE_FILE)
        return jsonify({'message': '状态更新成功'}), 200
    except Exception:
        logging.exception('更新版面费状态失败')
        return _server_error('状态更新失败')


@app.route('/save_page_fee_data', methods=['POST'])
def save_page_fee_data():
    try:
        updated_data = request.json
        wb = load_workbook(PAGE_FEE_FILE)
        ws = wb.active

        for item in updated_data:
            for row in ws.iter_rows(min_row=2):
                if row[1].value == item['稿件编号']:
                    row[0].value = item['备注']
                    row[2].value = item['核销号']
                    row[5].value = item['财务备注']
                    row[6].value = item['税号']
                    row[4].value = item['发票抬头']
                    row[7].value = item['邮箱']

                    if item.get('是否重复'):
                        for cell in row:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            cell.font = Font(color='9C0006')
                    break

        wb.save(PAGE_FEE_FILE)
        return jsonify({'message': '数据保存成功'})
    except Exception:
        logging.exception('保存版面费数据失败')
        return _server_error('保存数据失败')


@app.route('/start_scraping', methods=['POST'])
def start_scraping():
    year = request.json.get('year')
    issue = request.json.get('issue')
    if not year or not issue:
        return jsonify({'error': '缺少年份或期数'}), 400

    try:
        result = scrape_and_process_data(year, issue)
        return jsonify({'message': '爬虫任务完成', 'result': result}), 200
    except Exception:
        logging.exception('爬虫任务失败')
        return _server_error('爬虫任务失败')


@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory(app.static_folder, filename)


@app.route('/save_notes', methods=['POST'])
def save_notes():
    try:
        notes_data = request.json
        with open(NOTES_FILE, 'w', encoding='utf-8') as file:
            json.dump(notes_data, file, ensure_ascii=False, indent=4)
        return jsonify({'message': '笔记保存成功'}), 200
    except Exception:
        logging.exception('保存笔记失败')
        return _server_error('保存笔记失败')


@app.route('/load_notes', methods=['GET'])
def load_notes():
    try:
        if NOTES_FILE.exists():
            with open(NOTES_FILE, 'r', encoding='utf-8') as file:
                notes_data = json.load(file)
            return jsonify(notes_data), 200
        return jsonify({}), 200
    except Exception:
        logging.exception('加载笔记失败')
        return _server_error('加载笔记失败')


@app.route('/open_folder', methods=['GET'])
def open_folder():
    try:
        _open_directory(WORKSPACE_DIR)
        return jsonify({'message': '文件夹已打开'}), 200
    except Exception:
        logging.exception('打开工作文件夹失败')
        return _server_error('无法打开文件夹')


@app.route('/chat', methods=['POST'])
def chat():
    user_input = request.json.get('message')
    if not user_input:
        return jsonify({'error': '缺少消息内容'}), 400

    try:
        response = spark_chatbot.chat(user_input)
        return jsonify({'response': response}), 200
    except Exception:
        logging.exception('聊天请求失败')
        return _server_error('聊天失败')


@app.route('/open_program_folder', methods=['GET'])
def open_program_folder():
    try:
        _open_directory(PROGRAM_DIR)
        return jsonify({'message': '程序文件夹已打开'}), 200
    except Exception:
        logging.exception('打开程序文件夹失败')
        return _server_error('无法打开程序文件夹')


@app.route('/extract_authors', methods=['POST'])
def extract_authors():
    if 'pdf_file' not in request.files:
        return jsonify({'error': '没有文件'}), 400

    file = request.files['pdf_file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = UPLOAD_FOLDER / filename
        file.save(filepath)

        try:
            papers_info = extract_author_info(str(filepath))
            if papers_info:
                output_path = OUTPUT_PATH / f'{filename}_统计结果.xlsx'
                create_excel_report(papers_info, str(output_path))
                return jsonify(papers_info)
            return jsonify({'error': '未能提取到作者信息'}), 404
        except Exception:
            logging.exception('提取作者信息失败')
            return _server_error('处理文件时出错')

    return jsonify({'error': '不支持的文件类型'}), 400


@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.static_folder, 'images'), 'favicon.ico', mimetype='image/vnd.microsoft.icon')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


if __name__ == '__main__':
    debug_mode = os.getenv('FLASK_DEBUG', '').lower() == 'true'
    port = int(os.getenv('PORT', '5005'))
    app.run(debug=debug_mode, port=port)
