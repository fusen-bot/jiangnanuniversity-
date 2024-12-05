from flask import Flask, request, jsonify, send_from_directory
from data_processing import process_review_data, query_by_employee_id, query_by_name, query_by_manuscript_id_or_reviewer
from database import load_employee_data
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import logging
from openpyxl.utils import get_column_letter
from webpq import scrape_and_process_data
import json
import subprocess
import platform
from spark_chat_interactive import SparkChatBot

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

app = Flask(__name__, static_folder='../frontend')

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

# 设置文件路径
base_path = '/Users/changfusheng/Desktop/学报/PY_AUTO/data'
logging.info(f"基础路径设置为: {base_path}")

review_file = os.path.join(base_path, '评审.xls')
re_review_file = os.path.join(base_path, '复审.xls')
employee_file = os.path.join(base_path, '在职员工.xlsx')
retired_file = os.path.join(base_path, '退休员工.xlsx')

# 添加文件路径检查
for file_path in [review_file, re_review_file, employee_file, retired_file]:
    if os.path.exists(file_path):
        logging.info(f"文件存在: {file_path}")
    else:
        logging.error(f"文件不存在: {file_path}")

# 修改加载员工数据部分
try:
    logging.info("开始加载员工数据...")
    employee_data = load_employee_data(employee_file, retired_file)
    logging.info("员工数据加载成功")
except FileNotFoundError as e:
    logging.error(f"文件未找到: {str(e)}")
    # 可以选择设置一个默认值或者抛出异常
    employee_data = []
except Exception as e:
    logging.error(f"加载员工数据时发生未知错误: {str(e)}")
    raise

import numpy as np

# 添加输出路径配置
OUTPUT_PATH = '/Users/changfusheng/Desktop/学报/PY_AUTO/output'

# 设置版面费Excel文件路径
PAGE_FEE_FILE = '/Users/changfusheng/Library/CloudStorage/OneDrive-个人/文档/2024下.xlsx'

# 设置记事本数据文件路径
NOTES_FILE = '/Users/changfusheng/Desktop/学报/PY_AUTO/data/notes.json'

spark_chatbot = SparkChatBot()

@app.route('/process_review', methods=['POST'])
def process_review():
    target_month = request.json.get('target_month')
    if not target_month:
        return jsonify({"error": "缺少 target_month 参数"}), 400
    try:
        result = process_review_data(review_file, re_review_file, employee_file, retired_file, target_month)
        if isinstance(result, list) and result:
            # 将结果保存为 Excel 文件
            df = pd.DataFrame(result)
            output_file = os.path.join(OUTPUT_PATH, f'processed_review_{target_month}.xlsx')
            
            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            df.to_excel(output_file, index=False)
            return jsonify({"message": "数据处理成功", "file": output_file}), 200
        else:
            return jsonify({"error": "没有找到符合条件的数据"}), 404
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")  # 打印错误信息
        return jsonify({"error": f"处理失败：{str(e)}"}), 400

@app.route('/query_employee', methods=['GET'])
def query_employee():
    employee_id = request.args.get('employee_id')
    name = request.args.get('name')
    
    try:
        if employee_id:
            result = query_by_employee_id(employee_data, employee_id)
        elif name:
            result = query_by_name(employee_data, name)
        else:
            return jsonify({"error": "请提供员工ID或姓名"}), 400
        
        # 检查是否是错误消息
        if isinstance(result, dict) and "message" in result:
            return jsonify({"error": result["message"]}), 404
            
        # 确保返回格式统一
        formatted_result = {
            "type": "employee",
            "data": result  # result 现在总是一个列表
        }
        return jsonify(formatted_result)
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/query_manuscript', methods=['GET'])
def query_manuscript():
    query = request.args.get('query')
    result = query_by_manuscript_id_or_reviewer(review_file, re_review_file, query)
    
    # 格式化结果
    formatted_result = {
        "type": "manuscript",
        "data": result
    }
    return jsonify(formatted_result)

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


@app.route('/get_page_fee_data', methods=['GET'])
def get_page_fee_data():
    try:
        logging.debug(f"尝试读取件: {PAGE_FEE_FILE}")
        wb = load_workbook(PAGE_FEE_FILE)
        ws = wb.active

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:  # 确保稿件编号不为空
                status_value = row[10] if len(row) > 10 else ''
                statuses = set(status_value.split(', ')) if status_value and status_value != "已完成" else {'录用', '发票'}

                row_data = {
                    '备注': row[0] or '',
                    '稿件编号': str(row[1]),  # 直接返回完整的稿件编号
                    '核销号': row[2] or '',
                    '财务备注': row[5] or '' if len(row) > 5 else '',
                    '税号': row[6] or '' if len(row) > 6 else '',
                    '发票抬头': row[4] or '' if len(row) > 4 else '',
                    '邮箱': row[7] or '' if len(row) > 7 else '',
                    '录用': '录用' in statuses,
                    '发票': '发票' in statuses
                }
                data.append(row_data)

        data.reverse()
        logging.debug(f"读取到 {len(data)} 条数据")
        return jsonify(data)
    except Exception as e:
        logging.error(f"读取文件时发生错误: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/update_status', methods=['POST'])
def update_status():
    try:
        data = request.json
        manuscript_number = data.get('稿件编号')
        status_type = data.get('statusType')
        is_checked = data.get('isChecked')

        if not manuscript_number or status_type not in ['录用', '发票']:
            return jsonify({'error': '无效的请求数据'}), 400

        wb = load_workbook(PAGE_FEE_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[1].value == manuscript_number:  # B列是稿件编号
                status_cell = row[10]  # K列是状态列
                current_status = status_cell.value or ''
                statuses = set(current_status.split(', ')) if current_status else set()

                if is_checked:
                    statuses.add(status_type)
                else:
                    statuses.discard(status_type)

                if '录用' in statuses and '发票' in statuses:
                    status_cell.value = "已完成"
                else:
                    status_cell.value = ', '.join(sorted(statuses))
                break

        wb.save(PAGE_FEE_FILE)
        return jsonify({"message": "状态更新成功"}), 200
    except Exception as e:
        logging.error(f"更新状态时发生错误: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/save_page_fee_data', methods=['POST'])
def save_page_fee_data():
    try:
        updated_data = request.json
        
        # 读取Excel文件
        wb = load_workbook(PAGE_FEE_FILE)
        ws = wb.active
        
        # 更新Excel数据
        for item in updated_data:
            for row in ws.iter_rows(min_row=2):
                if row[1].value == item['稿件编号']:
                    row[0].value = item['备注']
                    row[2].value = item['核销号']
                    row[5].value = item['财务备注']
                    row[6].value = item['税号']
                    row[4].value = item['发票抬头']
                    row[7].value = item['邮箱']
                    
                    # 设置重复稿件的样式
                    if item.get('是否重复'):
                        for cell in row:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                    break
        
        # 保存更新后的Excel文件
        wb.save(PAGE_FEE_FILE)
        
        return jsonify({'message': '数据保存成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/start_scraping', methods=['POST'])
def start_scraping():
    year = request.json.get('year')
    issue = request.json.get('issue')
    if not year or not issue:
        return jsonify({"error": "缺少年份或期数"}), 400
    
    try:
        # 直接调用 scrape_and_process_data 函数
        result = scrape_and_process_data(year, issue)
        return jsonify({"message": "爬虫任务完成", "result": result}), 200
    except Exception as e:
        return jsonify({"error": f"爬虫任务失败: {str(e)}"}), 500

@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory(app.static_folder, filename)

@app.route('/save_notes', methods=['POST'])
def save_notes():
    try:
        notes_data = request.json
        with open(NOTES_FILE, 'w', encoding='utf-8') as f:
            json.dump(notes_data, f, ensure_ascii=False, indent=4)
        return jsonify({"message": "笔记保存成功"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/load_notes', methods=['GET'])
def load_notes():
    try:
        if os.path.exists(NOTES_FILE):
            with open(NOTES_FILE, 'r', encoding='utf-8') as f:
                notes_data = json.load(f)
            return jsonify(notes_data), 200
        else:
            return jsonify({}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/open_folder', methods=['GET'])
def open_folder():
    folder_path = '/Users/changfusheng/Desktop/学报'
    try:
        if platform.system() == "Darwin":  # macOS
            subprocess.run(["open", folder_path])
        elif platform.system() == "Windows":
            subprocess.run(["explorer", folder_path])
        else:  # Linux或其他系统
            subprocess.run(["xdg-open", folder_path])
        return jsonify({"message": "文件夹已打开"}), 200
    except Exception as e:
        return jsonify({"error": f"无法打开文件夹: {str(e)}"}), 500

@app.route('/chat', methods=['POST'])
def chat():
    user_input = request.json.get('message')
    if not user_input:
        print("错误：缺少消息内容")
        return jsonify({"error": "缺少消息内容"}), 400
    
    try:
        print(f"用户输入: {user_input}")
        response = spark_chatbot.chat(user_input)
        print(f"AI 响应: {response}")
        return jsonify({"response": response}), 200
    except Exception as e:
        error_message = f"聊天失败: {str(e)}"
        print(f"错误: {error_message}")
        return jsonify({"error": error_message}), 500

@app.route('/open_program_folder', methods=['GET'])
def open_program_folder():
    folder_path = '/Users/changfusheng/Desktop/学报/PY_AUTO'
    try:
        if platform.system() == "Darwin":  # macOS
            subprocess.run(["open", folder_path])
        elif platform.system() == "Windows":
            subprocess.run(["explorer", folder_path])
        else:  # Linux或其他系统
            subprocess.run(["xdg-open", folder_path])
        return jsonify({"message": "程序文件夹已打开"}), 200
    except Exception as e:
        return jsonify({"error": f"无法打开程序文件夹: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5005)
